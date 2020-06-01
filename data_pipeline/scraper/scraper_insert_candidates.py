
###
# NEEDS MORE WORK AND DATA CLARIFIATION
###


import pandas as pd
import os
import shutil
import xlrd
from openpyxl.workbook import Workbook

cwd = os.getcwd()
data_dir = cwd + '/Office_Sought_Backup'
filename = max([data_dir + "/" + f for f in os.listdir(data_dir)],key=os.path.getctime)

print(filename)
candidate_name = {'Cand_Nam L': 'TEST LAST', 'Cand_Nam F': 'TEST FIRST'}
candidate_data = {'Cand_Nam L': 0, 'Cand_Nam F': 0}

book = xlrd.open_workbook(filename)
sheet = book.sheet_by_index(0)
nrows = sheet.nrows
ncols = sheet.ncols

# prepare a xlsx sheet
book1 = Workbook()
sheet1 = book1.active

for row in range(1, nrows+1):
    for col in range(1, ncols+1):
        if row == 1:
            sheet1.cell(row=row, column=col).value = sheet.cell_value(row-1, col-1)
            for k, v in candidate_data.items():
                if k == sheet.cell_value(row-1, col-1):
                    candidate_data[k] = col-1
                    print(sheet.cell_value(row-1, col-1))
        else:
            sheet1.cell(row=row, column=col).value = sheet.cell_value(row-1, col-1)

for row in range(2, sheet1.max_row+1):
    for k, v in candidate_data.items():
        cell = sheet1.cell(row=row, column=v+1)
        cell.value = candidate_name[k]

book1.save('test.xlsx')

for k, v in candidate_data.items():
    print(pd.read_excel('test.xlsx', index_col=v, header=None))