import csv
import glob
import openpyxl
import xlrd

from open_workbook import open_workbook
from os import path, listdir, name
import pandas as pd
from time import sleep

# Custom python module
from dirmanager import DirManager

FILE_DIVIDER = '\\' if name == 'nt' else '/'

class PreProcessing():
    def __init__(self, scraper_download_dir):
        download_file_dir_wildcard = '{}/*.xls'.format(scraper_download_dir)
        self.filenames = glob.glob(download_file_dir_wildcard)
        self.download_dir = scraper_download_dir    
        self.insertCandidateFolder = DirManager(['insertedData'])
        self.insertCandidateFolder.createFolder()

    def aggregateData(self):
        # Create new directory for storing aggregated data
        # download_folder = path.abspath(path.join(self.download_dir, pardir))
        aggregateFolder = DirManager(['aggregated_data'])
        aggregateFolder.createFolder()
        new_folder = aggregateFolder.getDirectory()
        new_csv_file = '{}/data.csv'.format(new_folder)

        insertColumsFolder = self.insertCandidateFolder.getDirectory()
        filenames = sorted([insertColumsFolder + "/" + f for f in listdir(insertColumsFolder)], key=path.getmtime)

        with open(new_csv_file, 'w') as new_aggregate_csv:
            new_worksheet = csv.writer(new_aggregate_csv, quoting=csv.QUOTE_ALL)

            transactions = set()

            # Loop through all workbooks (EXCEL)
            header = False
            for filename in filenames:
                # Open worksheet
                wb = openpyxl.load_workbook(filename=filename)
                sheet = wb.worksheets[0]
                
                # Only pull excel header from the first file to reduce duplicates
                if not header:
                    new_worksheet.writerow(cell.value for cell in next(sheet.rows))
                # for rownum in range(1, sheet.max_row):
                #     # Skip duplicated entries.
                #     transaction_id = sheet[2]
                #     if transaction_id in transactions:
                #         continue
                #     transactions.add(transaction_id)
                #     new_worksheet.writerow(sheet.iter_rows(rownum+1, rownum+1))
                for row in sheet.iter_rows(2):
                    transaction_id = row[12].value
                    if transaction_id in transactions:
                        continue
                    transactions.add(transaction_id)
                    new_worksheet.writerow(cell.value for cell in row)

    def insertColumns(self, numDownloads, CandidateName, ElectionDate, BallotItem):
        print('Processing {} for {}'.format(numDownloads, CandidateName))

        if numDownloads == 0:
            return

        new_folder = self.insertCandidateFolder.getDirectory()
        filenames = self.insertColumnsHelper()


        candidateHeader = "CandidateControlledName"
        electionDateHeader = "Election Date"
        ballotItemHeader = "Ballot Item"

        # print(filenames)
        for fullfilepathname in filenames[-numDownloads:]:
            filename = path.basename(fullfilepathname)
            wb = open_workbook(fullfilepathname)
            if wb is None:
                continue
            errordTypes = ['Cmte_ID', 'Intr_Nam L', 'Intr_City', 'Intr_ST', 'Off_S_H_Cd', 'XRef_Match']
            data = pd.read_excel(wb, dtype={datatype: str for datatype in errordTypes})
            if CandidateName == "   ":
                data.insert(0, candidateHeader, "Independent")
            else:
                data.insert(0, candidateHeader, CandidateName)

            data.insert(0, electionDateHeader, ElectionDate)
            data.insert(0, ballotItemHeader, BallotItem)

            data.to_excel('{}/{}'.format(new_folder, filename + 'x'), index=False)
            
    
    def insertColumnsHelper(self):
        partial_download = True
        filenames = sorted([self.download_dir + FILE_DIVIDER + f for f in listdir(self.download_dir)], key=path.getmtime)

        while partial_download:
            filename = path.basename(filenames[-1])
            print(filename)
            if "transactionExportGrid" in filename and "crdownload" not in filename:
                partial_download = False
            else:
                sleep(3)
                filenames = sorted([self.download_dir + FILE_DIVIDER + f for f in listdir(self.download_dir)], key=path.getmtime)

        return filenames